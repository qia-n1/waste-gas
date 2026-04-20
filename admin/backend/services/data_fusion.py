from __future__ import annotations

import asyncio
import csv
import io
import json
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence

import httpx
from openpyxl import load_workbook

from config import settings
from services.vocs_proxy import SENSOR_FIELDS

# Global variable for time granularity control (minutes), easy for testing.
AGGREGATION_GRAIN_MINUTES = settings.aggregation_granularity_minutes

# Optional aliases for heterogeneous device keys.
FIELD_ALIASES: Dict[str, List[str]] = {
    "ambient_temp": ["temp", "temperature", "env_temp"],
    "ambient_humidity": ["humidity", "env_humidity"],
    "ambient_pressure": ["pressure", "env_pressure"],
    "rto_out_conc": ["voc", "vocs", "voc_conc", "rto_out_voc"],
    "rto_out_temp": ["rto_temp", "rto_out_temperature"],
}

SOURCE_FILE_MAP: Dict[str, str] = {
    "json": "json.csv",
    "jsonl": "json.csv",
    "txt": "json.csv",
    "csv": "csv.csv",
    "xlsx": "xlsx.csv",
    "xlsm": "xlsx.csv",
    "xltx": "xlsx.csv",
    "xltm": "xlsx.csv",
}


@dataclass
class FusionRuntimeState:
    running: bool = False
    last_aggregation_timestamp: str = ""
    last_history_upload_file: str = ""
    last_model_push_timestamp: str = ""
    last_model_push_ok: bool = False
    last_model_push_message: str = ""
    last_frontend_push_timestamp: str = ""
    last_frontend_push_ok: bool = False
    last_frontend_push_message: str = ""


_RUNTIME = FusionRuntimeState()
_TASK: Optional[asyncio.Task] = None
_LOCK = asyncio.Lock()


def _now() -> datetime:
    return datetime.now()


def _to_float(value: Any, default: float = 0.0) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _try_parse_timestamp(value: Any) -> Optional[datetime]:
    if isinstance(value, datetime):
        return value

    text = str(value or "").strip()
    if not text:
        return None

    if text.isdigit():
        raw = int(text)
        if raw > 1_000_000_000_000:
            return datetime.fromtimestamp(raw / 1000)
        return datetime.fromtimestamp(raw)

    for fmt in (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%dT%H:%M:%S%z",
    ):
        try:
            dt = datetime.strptime(text, fmt)
            if dt.tzinfo is not None:
                return dt.replace(tzinfo=None)
            return dt
        except ValueError:
            continue

    try:
        dt = datetime.fromisoformat(text.replace("Z", "+00:00"))
        if dt.tzinfo is not None:
            return dt.replace(tzinfo=None)
        return dt
    except ValueError:
        return None


def _parse_timestamp(value: Any) -> datetime:
    parsed = _try_parse_timestamp(value)
    if parsed is not None:
        return parsed
    return _now()


def _resolve_history_input_path(history_file: Optional[str] = None) -> Optional[Path]:
    _ensure_dirs()
    if not history_file:
        return None

    p = Path(history_file)
    if not p.is_absolute():
        p = settings.ingest_data_dir / p
    return p if p.exists() else None


def _normalize_aggregate_row(row: Dict[str, Any]) -> Dict[str, Any]:
    out: Dict[str, Any] = {"timestamp": str(row.get("timestamp", ""))}
    for field in SENSOR_FIELDS:
        raw = row.get(field)
        if raw in (None, ""):
            out[field] = 0.0
            continue
        try:
            out[field] = round(float(raw), 6)
        except (TypeError, ValueError):
            out[field] = 0.0
    return out


def _merge_aggregate_rows_incremental(new_rows: List[Dict[str, Any]]) -> Dict[str, int]:
    agg_path = _aggregate_csv_path()
    existing_rows = _load_csv_rows(agg_path)

    merged_by_ts: Dict[str, Dict[str, Any]] = {}
    for row in existing_rows:
        ts = str(row.get("timestamp", ""))
        if not ts:
            continue
        merged_by_ts[ts] = _normalize_aggregate_row(row)

    inserted = 0
    updated = 0
    for row in new_rows:
        ts = str(row.get("timestamp", ""))
        if not ts:
            continue
        norm = _normalize_aggregate_row(row)
        if ts in merged_by_ts:
            updated += 1
        else:
            inserted += 1
        merged_by_ts[ts] = norm

    sorted_rows = sorted(
        merged_by_ts.values(),
        key=lambda r: _parse_timestamp(r.get("timestamp")),
    )

    with agg_path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["timestamp", *SENSOR_FIELDS])
        writer.writeheader()
        for row in sorted_rows:
            writer.writerow(row)

    return {
        "inserted": inserted,
        "updated": updated,
        "total_after_merge": len(sorted_rows),
    }


def _rebuild_aggregate_from_history_csv(history_path: Path) -> Dict[str, Any]:
    rows = _load_csv_rows(history_path)
    if not rows:
        return {"ok": False, "message": f"{history_path.name} has no data", "written": 0}

    grain_seconds = max(1, AGGREGATION_GRAIN_MINUTES) * 60
    parsed_rows: List[tuple[datetime, Dict[str, Any]]] = []
    for row in rows:
        ts = _try_parse_timestamp(row.get("timestamp"))
        if ts is None:
            continue
        parsed_rows.append((ts, row))

    if not parsed_rows:
        return {"ok": False, "message": f"No valid timestamps in {history_path.name}", "written": 0}

    max_ts = max(item[0] for item in parsed_rows)
    buckets: Dict[int, Dict[str, List[float]]] = {}

    for ts, row in parsed_rows:
        delta_seconds = (max_ts - ts).total_seconds()
        if delta_seconds < 0:
            continue

        bucket_index = int(delta_seconds // grain_seconds)
        if bucket_index not in buckets:
            buckets[bucket_index] = {field: [] for field in SENSOR_FIELDS}

        for field in SENSOR_FIELDS:
            raw = row.get(field)
            if raw in (None, ""):
                continue
            try:
                buckets[bucket_index][field].append(float(raw))
            except (TypeError, ValueError):
                continue

    if not buckets:
        return {"ok": False, "message": f"No numeric sensor values in {history_path.name}", "written": 0}

    output_rows: List[Dict[str, Any]] = []
    for bucket_index in sorted(buckets.keys()):
        bucket_ts = (max_ts - timedelta(seconds=bucket_index * grain_seconds)).replace(second=0, microsecond=0)
        out_row: Dict[str, Any] = {
            "timestamp": bucket_ts.strftime("%Y-%m-%d %H:%M:00"),
        }
        for field in SENSOR_FIELDS:
            values = buckets[bucket_index][field]
            out_row[field] = round(sum(values) / len(values), 6) if values else 0.0
        output_rows.append(out_row)

    output_rows.sort(key=lambda row: row["timestamp"])
    merge_stats = _merge_aggregate_rows_incremental(output_rows)

    return {
        "ok": True,
        "message": f"Merged aggregate rows from {history_path.name}",
        "history_file": history_path.name,
        "written": len(output_rows),
        "inserted": merge_stats["inserted"],
        "updated": merge_stats["updated"],
        "aggregate_total": merge_stats["total_after_merge"],
        "source_rows": len(parsed_rows),
        "max_timestamp": max_ts.strftime("%Y-%m-%d %H:%M:%S"),
    }


async def rebuild_aggregate_from_history_csv(history_file: Optional[str] = None) -> Dict[str, Any]:
    async with _LOCK:
        if not history_file:
            return {
                "ok": False,
                "message": "history_file is required to avoid duplicate aggregation",
                "written": 0,
            }

        history_path = _resolve_history_input_path(history_file)
        if history_path is None:
            return {"ok": False, "message": "No history csv found", "written": 0}

        result = _rebuild_aggregate_from_history_csv(history_path)
        if result.get("ok") and result.get("written", 0) > 0:
            _RUNTIME.last_aggregation_timestamp = str(result.get("max_timestamp", ""))
            _RUNTIME.last_history_upload_file = str(result.get("history_file", ""))
        return result


def _ensure_dirs() -> None:
    settings.ingest_data_dir.mkdir(parents=True, exist_ok=True)


def _device_csv(device_id: str) -> Path:
    _ensure_dirs()
    safe = "".join(ch for ch in device_id if ch.isalnum() or ch in ("-", "_"))
    safe = safe or "unknown"
    return settings.ingest_data_dir / f"{safe}.csv"


def _source_csv(source_format: str) -> Path:
    _ensure_dirs()
    filename = SOURCE_FILE_MAP.get(source_format.lower())
    if not filename:
        raise ValueError(f"Unsupported source format: {source_format}")
    return settings.ingest_data_dir / filename


def _history_snapshot_csv(now: Optional[datetime] = None) -> Path:
    _ensure_dirs()
    stamp = (now or _now()).strftime("%Y%m%d_%H%M%S_%f")
    return settings.ingest_data_dir / f"device_history_{stamp}.csv"


def _aggregate_csv_path() -> Path:
    _ensure_dirs()
    return settings.ingest_data_dir / f"{AGGREGATION_GRAIN_MINUTES}.csv"


def _normalize_record(record: Dict[str, Any]) -> Dict[str, Any]:
    row: Dict[str, Any] = {}
    row["timestamp"] = _parse_timestamp(
        record.get("timestamp")
        or record.get("time")
        or record.get("ts")
    ).strftime("%Y-%m-%d %H:%M:00")

    for field in SENSOR_FIELDS:
        candidates = [field, *FIELD_ALIASES.get(field, [])]
        found = False
        for key in candidates:
            if key in record and record.get(key) not in (None, ""):
                row[field] = _to_float(record.get(key))
                found = True
                break
        if not found:
            # Keep sparse device record: missing fields stay blank in CSV.
            continue
    return row


def _extract_records(payload: Dict[str, Any]) -> List[Dict[str, Any]]:
    if "records" in payload and isinstance(payload["records"], list):
        return [item for item in payload["records"] if isinstance(item, dict)]
    if "data_sequence" in payload and isinstance(payload["data_sequence"], list):
        return [item for item in payload["data_sequence"] if isinstance(item, dict)]
    if "data" in payload and isinstance(payload["data"], dict):
        return [payload["data"]]
    if "record" in payload and isinstance(payload["record"], dict):
        return [payload["record"]]
    return [payload]


def _append_normalized_records(path: Path, normalized_rows: Sequence[Dict[str, Any]]) -> int:
    if not normalized_rows:
        return 0

    incoming_fields = set()
    for row in normalized_rows:
        for key, value in row.items():
            if key == "timestamp":
                continue
            if value in (None, ""):
                continue
            incoming_fields.add(key)

    existing_rows: List[Dict[str, Any]] = []
    existing_fields: List[str] = ["timestamp"]
    if path.exists():
        with path.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            if reader.fieldnames:
                existing_fields = [name for name in reader.fieldnames if name]
            existing_rows = [dict(row) for row in reader]

    merged_fields = ["timestamp"]
    for field in existing_fields:
        if field != "timestamp":
            merged_fields.append(field)
    for field in SENSOR_FIELDS:
        if field in incoming_fields and field not in merged_fields:
            merged_fields.append(field)

    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=merged_fields)
        writer.writeheader()
        for old_row in existing_rows:
            out = {"timestamp": old_row.get("timestamp", "")}
            for field in merged_fields:
                if field == "timestamp":
                    continue
                out[field] = old_row.get(field, "")
            writer.writerow(out)

        for row in normalized_rows:
            out = {"timestamp": row.get("timestamp", "")}
            for field in merged_fields:
                if field == "timestamp":
                    continue
                value = row.get(field, "")
                out[field] = "" if value in (None, "") else value
            writer.writerow(out)
    return len(normalized_rows)


def append_device_records(device_id: str, records: Sequence[Dict[str, Any]]) -> int:
    path = _device_csv(device_id)
    normalized_rows = [_normalize_record(rec) for rec in records]
    return _append_normalized_records(path, normalized_rows)


def append_source_records(source_format: str, records: Sequence[Dict[str, Any]]) -> tuple[int, Path]:
    path = _source_csv(source_format)
    normalized_rows = [_normalize_record(rec) for rec in records]
    return _append_normalized_records(path, normalized_rows), path


def replace_device_records(device_id: str, records: Sequence[Dict[str, Any]]) -> int:
    """Overwrite device csv with normalized records.

    This is used by history backfill uploads where each file represents
    a complete snapshot and should not be appended repeatedly.
    """
    path = _device_csv(device_id)
    normalized_rows = [_normalize_record(rec) for rec in records]
    if not normalized_rows:
        return 0

    fields = ["timestamp", *SENSOR_FIELDS]
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        for row in normalized_rows:
            out = {"timestamp": row.get("timestamp", "")}
            for field in SENSOR_FIELDS:
                value = row.get(field, "")
                out[field] = "" if value in (None, "") else value
            writer.writerow(out)
    return len(records)


def save_history_snapshot(records: Sequence[Dict[str, Any]]) -> tuple[int, Path]:
    snapshot_path = _history_snapshot_csv()
    normalized_rows = [_normalize_record(rec) for rec in records]
    if not normalized_rows:
        return 0, snapshot_path

    fields = ["timestamp", *SENSOR_FIELDS]
    with snapshot_path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        for row in normalized_rows:
            out = {"timestamp": row.get("timestamp", "")}
            for field in SENSOR_FIELDS:
                value = row.get(field, "")
                out[field] = "" if value in (None, "") else value
            writer.writerow(out)
    return len(records), snapshot_path


def parse_upload_file(content: bytes, filename: str) -> List[Dict[str, Any]]:
    suffix = Path(filename).suffix.lower()

    if suffix in {".json", ".jsonl", ".txt"}:
        text = content.decode("utf-8", errors="ignore").strip()
        if not text:
            return []
        if text.startswith("{"):
            try:
                obj = json.loads(text)
                if isinstance(obj, dict):
                    return _extract_records(obj)
            except json.JSONDecodeError:
                pass
        records: List[Dict[str, Any]] = []
        for line in text.splitlines():
            line = line.strip()
            if not line:
                continue
            parsed = json.loads(line)
            if isinstance(parsed, dict):
                records.append(parsed)
        return records

    if suffix == ".csv":
        text = content.decode("utf-8", errors="ignore")
        reader = csv.DictReader(io.StringIO(text))
        return [dict(row) for row in reader]

    if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        wb = load_workbook(filename=io.BytesIO(content), data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        records = []
        for r in rows[1:]:
            item: Dict[str, Any] = {}
            for idx, header in enumerate(headers):
                if not header:
                    continue
                item[header] = r[idx] if idx < len(r) else None
            if item:
                records.append(item)
        return records

    raise ValueError(f"Unsupported file type: {suffix}")


def _load_csv_rows(path: Path) -> List[Dict[str, Any]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return [dict(row) for row in csv.DictReader(f)]


def _window_average(now: datetime) -> Optional[Dict[str, Any]]:
    start = now - timedelta(minutes=AGGREGATION_GRAIN_MINUTES)
    values: Dict[str, List[float]] = {k: [] for k in SENSOR_FIELDS}

    for source_name in ("json.csv", "csv.csv", "xlsx.csv"):
        path = settings.ingest_data_dir / source_name
        if not path.exists():
            continue
        for row in _load_csv_rows(path):
            ts = _parse_timestamp(row.get("timestamp"))
            if ts <= start or ts > now:
                continue
            for field in SENSOR_FIELDS:
                raw = row.get(field)
                if raw in (None, ""):
                    continue
                try:
                    values[field].append(float(raw))
                except (TypeError, ValueError):
                    continue

    count = sum(1 for v in values.values() if v)
    if count == 0:
        return None

    out: Dict[str, Any] = {
        "timestamp": now.replace(second=0, microsecond=0).strftime("%Y-%m-%d %H:%M:00")
    }
    for field in SENSOR_FIELDS:
        field_vals = values[field]
        out[field] = round(sum(field_vals) / len(field_vals), 6) if field_vals else 0.0
    return out


def append_aggregate_row(row: Dict[str, Any]) -> None:
    path = _aggregate_csv_path()
    fields = ["timestamp", *SENSOR_FIELDS]
    exists = path.exists()
    with path.open("a", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        if not exists:
            writer.writeheader()
        writer.writerow(row)


def _build_predict_payload() -> Optional[Dict[str, Any]]:
    rows = _load_csv_rows(_aggregate_csv_path())
    if not rows:
        return None

    # Partial-device ingestion may produce sparse aggregate fields in early stages.
    # For cold start, left-pad with zero-valued rows so the model still receives a 96-step sequence.
    last_seen: Dict[str, float] = {field: 0.0 for field in SENSOR_FIELDS}
    data_sequence = []
    trimmed_rows = rows[-96:]
    pad_count = max(0, 96 - len(trimmed_rows))

    first_timestamp = _parse_timestamp(trimmed_rows[0].get("timestamp")) if trimmed_rows else _now()
    pad_start = first_timestamp - timedelta(minutes=AGGREGATION_GRAIN_MINUTES * pad_count)

    for pad_index in range(pad_count):
        data_sequence.append(
            {
                "timestamp": (pad_start + timedelta(minutes=AGGREGATION_GRAIN_MINUTES * pad_index)).strftime("%Y-%m-%d %H:%M:00"),
                "feature_values": [0.0 for _ in SENSOR_FIELDS],
            }
        )

    for row in trimmed_rows:
        feature_values: List[float] = []
        for field in SENSOR_FIELDS:
            raw = row.get(field)
            if raw in (None, ""):
                feature_values.append(last_seen[field])
                continue
            try:
                val = float(raw)
                last_seen[field] = val
                feature_values.append(val)
            except (TypeError, ValueError):
                feature_values.append(last_seen[field])

        data_sequence.append(
            {
                "timestamp": str(row.get("timestamp", "")),
                "feature_values": feature_values,
            }
        )
    return {"data_sequence": data_sequence}


async def _send_to_model(payload: Dict[str, Any]) -> Dict[str, Any]:
    async with httpx.AsyncClient(base_url=settings.vocs_base_url, timeout=settings.request_timeout) as client:
        response = await client.post("/predict", json=payload)
        return {
            "ok": response.status_code == 200,
            "status_code": response.status_code,
            "body": response.json() if response.content else {},
        }


async def _push_to_frontend(payload: Dict[str, Any]) -> Dict[str, Any]:
    url = settings.frontend_push_url.strip()
    if not url:
        return {
            "ok": False,
            "status_code": 0,
            "message": "FRONTEND_PUSH_URL not configured",
        }

    async with httpx.AsyncClient(timeout=settings.frontend_push_timeout) as client:
        response = await client.post(url, json=payload)
        return {
            "ok": response.status_code == 200,
            "status_code": response.status_code,
            "body": response.json() if response.content else {},
        }


async def run_aggregate_and_push_once() -> Dict[str, Any]:
    async with _LOCK:
        now = _now()
        row = _window_average(now)
        if row is None:
            return {"ok": False, "message": "no rows in current window"}

        append_aggregate_row(row)
        _RUNTIME.last_aggregation_timestamp = row["timestamp"]

        payload = _build_predict_payload()
        if payload is None:
            msg = "no aggregated rows available for model push"
            _RUNTIME.last_model_push_ok = False
            _RUNTIME.last_model_push_message = msg
            return {"ok": False, "message": msg, "aggregated_at": row["timestamp"]}

        try:
            result = await _send_to_model(payload)

            _RUNTIME.last_model_push_timestamp = _now().isoformat()
            _RUNTIME.last_model_push_ok = bool(result.get("ok"))
            _RUNTIME.last_model_push_message = f"status={result.get('status_code')}"

            frontend_push_result: Dict[str, Any] | None = None
            if result.get("ok"):
                frontend_payload = {
                    "event": "data-fusion-model-result",
                    "aggregated_at": row["timestamp"],
                    "model_result": result,
                    "sent_at": _now().isoformat(),
                }
                frontend_push_result = await _push_to_frontend(frontend_payload)
                _RUNTIME.last_frontend_push_timestamp = _now().isoformat()
                _RUNTIME.last_frontend_push_ok = bool(frontend_push_result.get("ok"))
                _RUNTIME.last_frontend_push_message = (
                    f"status={frontend_push_result.get('status_code')}"
                    if settings.frontend_push_url.strip()
                    else "FRONTEND_PUSH_URL not configured"
                )

            return {
                "ok": bool(result.get("ok")),
                "aggregated_at": row["timestamp"],
                "model_result": result,
                "frontend_push": frontend_push_result,
            }
        except Exception as exc:  # noqa: BLE001
            _RUNTIME.last_model_push_timestamp = _now().isoformat()
            _RUNTIME.last_model_push_ok = False
            _RUNTIME.last_model_push_message = str(exc)
            return {
                "ok": False,
                "aggregated_at": row["timestamp"],
                "message": f"model push failed: {exc}",
            }


async def _scheduler_loop() -> None:
    while _RUNTIME.running:
        await run_aggregate_and_push_once()
        await asyncio.sleep(max(1, AGGREGATION_GRAIN_MINUTES) * 60)


def start_scheduler() -> None:
    global _TASK
    if _RUNTIME.running:
        return
    _ensure_dirs()
    _RUNTIME.running = True
    _TASK = asyncio.create_task(_scheduler_loop())


async def stop_scheduler() -> None:
    global _TASK
    _RUNTIME.running = False
    if _TASK is not None:
        _TASK.cancel()
        try:
            await _TASK
        except asyncio.CancelledError:
            pass
        _TASK = None


def get_status() -> Dict[str, Any]:
    files = sorted([p.name for p in settings.ingest_data_dir.glob("*.csv")]) if settings.ingest_data_dir.exists() else []
    return {
        "running": _RUNTIME.running,
        "aggregation_granularity_minutes": AGGREGATION_GRAIN_MINUTES,
        "ingest_data_dir": str(settings.ingest_data_dir),
        "aggregate_csv": str(_aggregate_csv_path()),
        "csv_files": files,
        "last_aggregation_timestamp": _RUNTIME.last_aggregation_timestamp,
        "last_history_upload_file": _RUNTIME.last_history_upload_file,
        "last_model_push_timestamp": _RUNTIME.last_model_push_timestamp,
        "last_model_push_ok": _RUNTIME.last_model_push_ok,
        "last_model_push_message": _RUNTIME.last_model_push_message,
        "last_frontend_push_timestamp": _RUNTIME.last_frontend_push_timestamp,
        "last_frontend_push_ok": _RUNTIME.last_frontend_push_ok,
        "last_frontend_push_message": _RUNTIME.last_frontend_push_message,
    }
